from openpyxl import load_workbook #o que é load_workbook?

# 1- Lê pasta de trabalho e polanilha
wb = load_workbook('data/pivot_table.xlsx') #carrego o workbook
sheet = wb['Relatório'] #indico o dado e a planilha que quero usar

# 2- Acessando um valor específico
print(sheet['A3'].value) #peguei aqui um valor específico da tabela pivot , LINHA ANO
print(sheet['B3'].value) 

# 3- Iterando valores por meio de loop
for i in range(2, 6):
    ano = sheet['A%s'%i].value #porque entram esses caracteres de percent?
    print(ano)
    
for i in range(2,6):
    ano = sheet['A%s' %i].value
    am = sheet['B%s' %i].value
    bt = sheet['C%s' %i].value
    print('{0} o Aston Martin vendeu {1} e o Bentley vendeu {2}'.format(ano, am, bt)) #Como aqui funciona?
    
