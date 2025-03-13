from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference # para criar um gráfico de barras

# 1- Lê pasta de trabalho e polanilha
wb = load_workbook('data/pivot_table.xlsx')
sheet = wb['Relatório']


# 2- Referências das Linhas e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
print(min_column)
print(max_column)
min_row = wb.active.min_row
max_row = wb.active.max_row


#3- Adicionando Dados e categorias no Gráfico
barchart = BarChart()

data = Reference(
    sheet,
    min_col=min_column + 1,
    max_col=max_column,
    min_row=min_row,
    max_row=max_row  # ✅ Corrigido (sem espaço)
)

categories = Reference(
    sheet,
    min_col=min_column,
    max_col=min_column,
    min_row=min_row + 1,
    max_row=max_row  # ✅ Corrigido (sem espaço)
)

barchart.add_data(data, titles_from_data=True) #pego as informações e passo para o gráfico
barchart.set_categories(categories)

# 4 - Crinado o Gráfico
sheet.add_chart(barchart, 'B10')
barchart.title = 'Vendas por Fabricantes'
barchart.style = 2

# Salvando o Workbook
wb.save('data/barchart.xlsx')
