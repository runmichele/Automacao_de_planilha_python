from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 

# 1- Lê pasta de trabalho e polanilha
wb = load_workbook('data/barchart.xlsx')
sheet = wb['Relatório']


# 2- Referências das Linhas e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
print(min_column)
print(max_column)
min_row = wb.active.min_row
max_row = wb.active.max_row


#3- Incluindo fórmula
# sheet['B6'] = '=SUM(B2:B5)' #como funciona esse cógigo?
# sheet['B6'].style = 'Currency'
for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i) 
    print(letter)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'


# Salva o arquivo na mesma pasta onde o script está rodando
wb.save('data/arquivo.xlsx')  # Agora será salvo dentro da pasta "data"
wb.save('data/test.xlsx')
